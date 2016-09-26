from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import time
import backend
import SpotipyManager


dest_filename = "test.xlsx"
try:
    wb = load_workbook(filename = dest_filename)
except:
    wb = Workbook()

ws = wb.active

indexread = False
index = 0

def setupExcelFormatting():
    print("Preparing the excel file")
    ws['A1'] = "Index:"
    ws['B1'] = "Artist:"
    ws['C1'] = "Songname:"
    ws['D1'] = "Started playing:"
    ws['E1'] = "Ended playing:"

def eraseExcelData():
    '''
    Clear all cells starting from third row. we keep our header and
    replaces other rows content by None
    '''
    print("Erasing all data")
    for row in ws.iter_rows(row_offset=0):
        for cell in row:
            cell.value=None
    wb.save(dest_filename)
    print("Erased everything!")

def writeNewSongToFile(songname):
    #Here we write everything to the Excel sheet
    startTime = str(datetime.datetime.now())
    print("Song started at " +startTime)

    song, artist = backend.getSongData(songname)
    SpotipyManager.searchSong(song,artist)

    ws.cell(row=index, column=2).value = artist
    ws.cell(row=index, column=3).value = song
    ws.cell(row=index, column=4).value = startTime
    if index > 2: #Notating the endtime
        ws.cell(row=index-1, column=5).value = startTime

    #EVerything committed now push!
    print("Saving Sheet...")
    wb.save(dest_filename)
    print("Saved!")

def writeIndex(index):
    ws['A2'] = index #first update the index of number of played songs ever

def manageIndex(songname):
    global indexread
    global index

    if indexread == False:
        index = ws['A2'].value
        indexread = True #we have now read index from the excel file
        print("index read from excel file = " + str(index))
        if index == None or index == 0:
            index = 2
            print("index = none, thus index = 2")
    elif songname != "Spotify" and songname != "":  # if we have already read an index from the excel file we add one to the index
        index += 1

    #Now lets write the new song to the excel file
    writeIndex(index)
